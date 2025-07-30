# --- cfdi_processor/excel_exporter.py ---
import pandas as pd
import os
# Importar para el autoajuste de ancho de columna
from openpyxl.utils import get_column_letter
# Importar órdenes de columna
from constants import INVOICE_COLUMN_ORDER, PAGOS_COLUMN_ORDER


def export_to_excel(invoice_data_list, nomina_data_list, pagos_data_list, output_file_path):
    """
    Exporta listas de diccionarios (una para facturas, otra para nóminas, otra para pagos)
    a un archivo de Excel con hojas separadas usando Pandas.

    Args:
        invoice_data_list (list): Una lista de diccionarios para facturas regulares.
        nomina_data_list (list): Lista de diccionarios para el complemento de nómina.
        pagos_data_list (list): Lista de diccionarios para el complemento de pagos.
        output_file_path (str): La ruta completa donde se guardará el archivo de Excel.
    """
    if not invoice_data_list and not nomina_data_list and not pagos_data_list:
        print("No hay datos para exportar. No se creará el archivo de Excel.")
        return

    # Asegurarse de que el directorio de salida exista.
    output_dir = os.path.dirname(output_file_path)
    # Si output_file_path era solo un nombre de archivo (ej., "informe.xlsx"), output_dir está vacío
    if not output_dir:
        # Fallback a una carpeta de informes predeterminada en el directorio de trabajo actual
        output_dir = os.path.join(os.getcwd(), "CFDI_Processor_App", "Reports")
    os.makedirs(output_dir, exist_ok=True)

    try:
        with pd.ExcelWriter(output_file_path, engine='openpyxl') as writer:
            # --- Hoja de Invoices ---
            if invoice_data_list:
                df_invoices = pd.DataFrame(invoice_data_list)

                # Reindexar el DataFrame para que coincida exactamente con el orden de columnas deseado.
                # Esto añadirá columnas faltantes con NaN y eliminará las no especificadas.
                # Excluimos 'CFDI_Type' ya que es una columna interna para categorización.
                final_invoice_columns = [
                    col for col in INVOICE_COLUMN_ORDER if col != "CFDI_Type"]
                df_invoices = df_invoices.reindex(
                    columns=final_invoice_columns)

                df_invoices.to_excel(
                    writer, sheet_name='Invoices', index=False)

                # Auto-ajustar el ancho de las columnas
                worksheet = writer.sheets['Invoices']
                for i, col in enumerate(df_invoices.columns):
                    max_length = 0
                    # Considerar la longitud del encabezado
                    max_length = max(max_length, len(str(col)))
                    # Iterar a través de la columna para encontrar la longitud máxima del contenido de la celda
                    for cell in worksheet.iter_cols(min_col=i+1, max_col=i+1, min_row=1):
                        for c in cell:
                            try:
                                if c.value is not None:
                                    # Convertir a cadena para medir la longitud, pero no cambiar el tipo de dato subyacente
                                    cell_value_str = str(c.value)
                                    # Si es un número flotante, redondear para la medición de longitud
                                    if isinstance(c.value, (float, int)):
                                        # Considerar 2 decimales para ancho
                                        cell_value_str = f"{c.value:.2f}"
                                    max_length = max(
                                        max_length, len(cell_value_str))
                            except TypeError:
                                pass  # Manejar casos donde el valor de la celda es None o no es una cadena
                    adjusted_width = (max_length + 2)
                    worksheet.column_dimensions[get_column_letter(
                        i + 1)].width = adjusted_width

                print(
                    f"Exportadas {len(invoice_data_list)} facturas CFDI regulares a la hoja 'Invoices'.")
            else:
                print("No hay datos de Facturas para exportar.")

            # --- Hoja de Nomina ---
            if nomina_data_list:
                df_nominas = pd.DataFrame(nomina_data_list)

                # Para la hoja de Nómina, no tenemos un orden estricto en constants.py,
                # así que simplemente eliminamos la columna interna 'CFDI_Type'.
                df_nominas = df_nominas.drop(
                    columns=['CFDI_Type'], errors='ignore')

                df_nominas.to_excel(writer, sheet_name='Nomina', index=False)

                # Auto-ajustar el ancho de las columnas
                worksheet = writer.sheets['Nomina']
                for i, col in enumerate(df_nominas.columns):
                    max_length = 0
                    max_length = max(max_length, len(str(col)))
                    for cell in worksheet.iter_cols(min_col=i+1, max_col=i+1, min_row=1):
                        for c in cell:
                            try:
                                if c.value is not None:
                                    cell_value_str = str(c.value)
                                    if isinstance(c.value, (float, int)):
                                        cell_value_str = f"{c.value:.2f}"
                                    max_length = max(
                                        max_length, len(cell_value_str))
                            except TypeError:
                                pass
                    adjusted_width = (max_length + 2)
                    worksheet.column_dimensions[get_column_letter(
                        i + 1)].width = adjusted_width

                print(
                    f"Exportados {len(nomina_data_list)} complementos de Nómina CFDI 1.2 a la hoja 'Nomina'.")
            else:
                print("No hay datos de Nómina para exportar.")

            # --- Hoja de Pagos ---
            if pagos_data_list:
                df_pagos = pd.DataFrame(pagos_data_list)

                # Reindexar el DataFrame para que coincida exactamente con el orden de columnas deseado.
                # Excluimos 'CFDI_Type' ya que es una columna interna para categorización.
                final_pagos_columns = [
                    col for col in PAGOS_COLUMN_ORDER if col != "CFDI_Type"]
                df_pagos = df_pagos.reindex(columns=final_pagos_columns)

                df_pagos.to_excel(writer, sheet_name='Pagos', index=False)

                # Auto-ajustar el ancho de las columnas
                worksheet = writer.sheets['Pagos']
                for i, col in enumerate(df_pagos.columns):
                    max_length = 0
                    max_length = max(max_length, len(str(col)))
                    for cell in worksheet.iter_cols(min_col=i+1, max_col=i+1, min_row=1):
                        for c in cell:
                            try:
                                if c.value is not None:
                                    cell_value_str = str(c.value)
                                    if isinstance(c.value, (float, int)):
                                        cell_value_str = f"{c.value:.2f}"
                                    max_length = max(
                                        max_length, len(cell_value_str))
                            except TypeError:
                                pass
                    adjusted_width = (max_length + 2)
                    worksheet.column_dimensions[get_column_letter(
                        i + 1)].width = adjusted_width

                print(
                    f"Exportados {len(pagos_data_list)} complementos de Pagos CFDI 2.0 a la hoja 'Pagos'.")
            else:
                print("No hay datos de Pagos para exportar.")

        print(f"\nDatos exportados exitosamente a Excel: {output_file_path}")

    except Exception as e:
        print(f"Error al exportar a Excel: {e}")
        print("Por favor, asegúrate de que 'openpyxl' esté instalado (pip install openpyxl) y la ruta de salida sea válida.")
